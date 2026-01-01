"""
Document Processing Module for GRC Policy Files
Handles PDF, Excel, TXT, and DOCX policy documents
"""

import io
import re
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
import PyPDF2
import openpyxl
from openpyxl import load_workbook


@dataclass
class DocumentSection:
    """Represents a section of a document"""
    section_id: str
    title: str
    content: str
    level: int
    page_number: Optional[int] = None
    subsections: List['DocumentSection'] = None
    
    def __post_init__(self):
        if self.subsections is None:
            self.subsections = []


class PDFPolicyProcessor:
    """Process PDF policy documents"""
    
    def __init__(self):
        self.metadata: Dict[str, Any] = {}
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text_content = []
        
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                # Extract metadata
                self.metadata = {
                    "pages": len(pdf_reader.pages),
                    "title": pdf_reader.metadata.get('/Title', 'Unknown') if pdf_reader.metadata else 'Unknown'
                }
                
                # Extract text from each page
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    text_content.append(f"[Page {page_num}]\n{text}\n")
        
        except Exception as e:
            raise Exception(f"Error processing PDF: {str(e)}")
        
        return "\n".join(text_content)
    
    def extract_structured_content(self, file_path: str) -> List[DocumentSection]:
        """Extract structured sections from PDF"""
        text = self.extract_text(file_path)
        return self._parse_sections(text)
    
    def _parse_sections(self, text: str) -> List[DocumentSection]:
        """Parse text into structured sections"""
        sections = []
        current_section = None
        section_counter = 0
        
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Detect section headers (various formats)
            # Format: "1. SECTION TITLE" or "SECTION TITLE" (all caps)
            section_match = re.match(r'^(\d+\.?\d*\.?)\s+([A-Z][A-Z\s]{3,})', line)
            if section_match or re.match(r'^[A-Z\s]{10,}$', line):
                if current_section:
                    sections.append(current_section)
                
                section_counter += 1
                current_section = DocumentSection(
                    section_id=f"S{section_counter:03d}",
                    title=line,
                    content="",
                    level=1
                )
            elif current_section:
                current_section.content += line + "\n"
        
        if current_section:
            sections.append(current_section)
        
        return sections


class ExcelPolicyProcessor:
    """Process Excel policy documents"""
    
    def __init__(self):
        self.metadata: Dict[str, Any] = {}
    
    def extract_policies(self, file_path: str) -> Dict[str, Any]:
        """Extract policy data from Excel file"""
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            self.metadata = {
                "sheets": workbook.sheetnames,
                "total_sheets": len(workbook.sheetnames)
            }
            
            policies = {
                "metadata": self.metadata,
                "sheets": {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                policies["sheets"][sheet_name] = self._process_sheet(sheet)
            
            workbook.close()
            return policies
        
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
    
    def _process_sheet(self, sheet) -> Dict[str, Any]:
        """Process a single worksheet"""
        
        data = {
            "headers": [],
            "rows": [],
            "policy_rules": []
        }
        
        rows_iter = sheet.iter_rows(values_only=True)
        
        # Get headers
        headers = next(rows_iter, None)
        if headers:
            data["headers"] = [h if h else f"Column_{i}" for i, h in enumerate(headers)]
        
        # Process data rows
        for row in rows_iter:
            if any(cell for cell in row):  # Skip empty rows
                row_data = dict(zip(data["headers"], row))
                data["rows"].append(row_data)
                
                # Try to extract policy rules
                policy_rule = self._extract_policy_from_row(row_data, data["headers"])
                if policy_rule:
                    data["policy_rules"].append(policy_rule)
        
        return data
    
    def _extract_policy_from_row(self, row_data: Dict[str, Any], headers: List[str]) -> Optional[Dict[str, Any]]:
        """Extract policy rule from Excel row"""
        
        # Common Excel policy formats:
        # - Policy ID | Policy Name | Description | Requirement | Category
        # - Rule ID | Description | Compliance Level | Requirements
        
        policy_columns = ['policy', 'rule', 'requirement', 'description']
        
        # Check if row contains policy information
        has_policy = any(
            any(col_keyword in str(header).lower() for col_keyword in policy_columns)
            for header in headers
        )
        
        if not has_policy:
            return None
        
        policy_rule = {}
        
        for header, value in row_data.items():
            if value:
                header_lower = str(header).lower()
                
                if 'policy' in header_lower or 'rule' in header_lower:
                    policy_rule['rule_name'] = str(value)
                elif 'description' in header_lower:
                    policy_rule['description'] = str(value)
                elif 'requirement' in header_lower:
                    policy_rule['requirement'] = str(value)
                elif 'category' in header_lower or 'type' in header_lower:
                    policy_rule['category'] = str(value)
                elif 'level' in header_lower or 'priority' in header_lower:
                    policy_rule['compliance_level'] = str(value)
        
        return policy_rule if policy_rule else None
    
    def convert_to_text(self, file_path: str) -> str:
        """Convert Excel content to text format"""
        
        policies = self.extract_policies(file_path)
        text_parts = []
        
        for sheet_name, sheet_data in policies["sheets"].items():
            text_parts.append(f"\n{'='*60}\nSHEET: {sheet_name}\n{'='*60}\n")
            
            # If policy rules were found, format them
            if sheet_data["policy_rules"]:
                for i, rule in enumerate(sheet_data["policy_rules"], 1):
                    text_parts.append(f"\nPolicy Rule {i}:")
                    for key, value in rule.items():
                        text_parts.append(f"  {key}: {value}")
            else:
                # Otherwise, format as table
                for row in sheet_data["rows"]:
                    row_text = " | ".join(str(v) if v else "" for v in row.values())
                    text_parts.append(row_text)
        
        return "\n".join(text_parts)


class TextPolicyProcessor:
    """Process plain text policy documents"""
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from file"""
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()
    
    def extract_structured_content(self, file_path: str) -> List[DocumentSection]:
        """Extract structured sections"""
        text = self.extract_text(file_path)
        return self._parse_sections(text)
    
    def _parse_sections(self, text: str) -> List[DocumentSection]:
        """Parse text into sections"""
        sections = []
        current_section = None
        section_counter = 0
        
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Detect section headers
            if self._is_section_header(line):
                if current_section:
                    sections.append(current_section)
                
                section_counter += 1
                current_section = DocumentSection(
                    section_id=f"S{section_counter:03d}",
                    title=line,
                    content="",
                    level=self._get_section_level(line)
                )
            elif current_section:
                current_section.content += line + "\n"
        
        if current_section:
            sections.append(current_section)
        
        return sections
    
    def _is_section_header(self, line: str) -> bool:
        """Check if line is a section header"""
        
        # Various header patterns
        patterns = [
            r'^\d+\.?\s+[A-Z]',  # "1. SECTION" or "1 SECTION"
            r'^[A-Z\s]{10,}$',    # All caps, 10+ chars
            r'^[A-Z][a-z]+(\s+[A-Z][a-z]+){2,}',  # Title Case (3+ words)
        ]
        
        return any(re.match(pattern, line) for pattern in patterns)
    
    def _get_section_level(self, line: str) -> int:
        """Determine section nesting level"""
        
        # Count leading numbering depth (1.2.3 -> level 3)
        number_match = re.match(r'^([\d\.]+)', line)
        if number_match:
            return number_match.group(1).count('.')
        
        return 1


class DocumentProcessorFactory:
    """Factory for creating document processors"""
    
    @staticmethod
    def get_processor(file_type: str):
        """Get appropriate processor for file type"""
        
        file_type = file_type.lower()
        
        if file_type in ['pdf']:
            return PDFPolicyProcessor()
        elif file_type in ['xlsx', 'xls', 'excel']:
            return ExcelPolicyProcessor()
        elif file_type in ['txt', 'text']:
            return TextPolicyProcessor()
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
    
    @staticmethod
    def process_file(file_path: str, file_type: str) -> str:
        """Process file and return text content"""
        
        processor = DocumentProcessorFactory.get_processor(file_type)
        
        if isinstance(processor, PDFPolicyProcessor):
            return processor.extract_text(file_path)
        elif isinstance(processor, ExcelPolicyProcessor):
            return processor.convert_to_text(file_path)
        elif isinstance(processor, TextPolicyProcessor):
            return processor.extract_text(file_path)
        else:
            raise ValueError(f"Unknown processor type")


# Testing utilities
def test_processors():
    """Test document processors"""
    
    # Test text processor
    sample_text = """
    CORPORATE DATA GOVERNANCE POLICY
    
    1. INTRODUCTION
    
    This policy establishes requirements for data governance.
    
    2. DATA CLASSIFICATION
    
    2.1 All data must be classified as Public, Internal, Confidential, or Restricted.
    
    2.2 Restricted data shall be encrypted using AES-256.
    
    3. ACCESS CONTROL
    
    3.1 Access to confidential data requires manager approval.
    """
    
    # Save to file
    with open('/tmp/test_policy.txt', 'w') as f:
        f.write(sample_text)
    
    # Process
    processor = TextPolicyProcessor()
    sections = processor.extract_structured_content('/tmp/test_policy.txt')
    
    print(f"Extracted {len(sections)} sections:")
    for section in sections:
        print(f"  - {section.title}")
        print(f"    Content preview: {section.content[:100]}...")


if __name__ == "__main__":
    test_processors()
